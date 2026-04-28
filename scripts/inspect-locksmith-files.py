from pathlib import Path
import json
import sys

from openpyxl import load_workbook

paths = [Path(p) for p in sys.argv[1:]]
summary = []

for path in paths:
    item = {"path": str(path), "exists": path.exists(), "type": path.suffix.lower()}
    if path.exists() and path.suffix.lower() in {".xlsx", ".xlsm"}:
        wb = load_workbook(path, read_only=True, data_only=True)
        sheets = []
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8), values_only=True):
                rows.append([cell for cell in row[:12]])
            sheets.append(
                {
                    "name": ws.title,
                    "maxRow": ws.max_row,
                    "maxColumn": ws.max_column,
                    "sampleRows": rows,
                }
            )
        item["sheets"] = sheets
    summary.append(item)

print(json.dumps(summary, indent=2, default=str))
