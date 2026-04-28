from pathlib import Path
import json
import sys

from openpyxl import load_workbook

source = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(r"C:\Users\th018\OneDrive\Desktop\Locksmith_Master_Catalog_v16.xlsx")
out_dir = Path("data")
out_dir.mkdir(exist_ok=True)
output = out_dir / "master-catalog.json"

wb = load_workbook(source, read_only=True, data_only=True)
ws = wb["Catalog"]
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
rows = []

def clean(value):
    if value is None:
        return ""
    return str(value).strip()

def split_pipe(value):
    return [part.strip() for part in clean(value).split("|") if part.strip()]

for row in ws.iter_rows(min_row=2, values_only=True):
    item = dict(zip(headers, row))
    hl_part = clean(item.get("HL Part #"))
    fcc = clean(item.get("FCC ID"))
    attrs = clean(item.get("Attributes"))
    oem = split_pipe(item.get("OEM Part Numbers"))
    mw = clean(item.get("MW P/N (Legacy)"))
    lr = clean(item.get("LR P/N (Legacy)"))
    ti = clean(item.get("TI P/N (Active)"))
    klr = clean(item.get("KLR P/N (Active)"))
    notes = clean(item.get("Notes"))
    if not any([hl_part, fcc, attrs, oem, mw, lr, ti, klr, notes]):
        continue
    rows.append(
        {
            "hlPartNumber": hl_part,
            "fccId": fcc,
            "attributes": attrs,
            "oemPartNumbers": oem,
            "mwLegacyPartNumber": mw,
            "lrLegacyPartNumber": lr,
            "tiActivePartNumber": ti,
            "klrActivePartNumber": klr,
            "notes": notes,
            "source": source.name,
        }
    )

payload = {
    "generatedAt": __import__("datetime").datetime.utcnow().isoformat() + "Z",
    "source": str(source),
    "totalRows": len(rows),
    "rows": rows,
}

output.write_text(json.dumps(payload, indent=2), encoding="utf-8")
print(json.dumps({"totalRows": len(rows), "outputPath": str(output), "sample": rows[:5]}, indent=2))
